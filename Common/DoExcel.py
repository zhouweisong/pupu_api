"""
-------------------------------------------------
   File Name：     DuExcel
   Description :
   Author :       zws
   date：          2018/3/13
-------------------------------------------------
   Change Activity:
                   2018/3/13:
-------------------------------------------------
"""
__author__ = 'zws'

from openpyxl import load_workbook
from Common.myLogger import *

class DoExcel:

    def __init__(self,excelpath):
        self.wb = load_workbook(excelpath)
        self.excelpath = excelpath
        self.sh_case_data = self.wb.get_sheet_by_name("case_datas")
        self.sh_prepar_data = self.wb.get_sheet_by_name("prepare_datas")

    #从excel中prepar_data读取初始化数据
    def get_init_datas(self):
        init_datas = {}
        for index in range(2, self.sh_prepar_data.max_row + 1):
            key = self.sh_prepar_data.cell(row=index,column =1).value
            init_datas[key] = self.sh_prepar_data.cell(row=index,column =2).value
            init_datas["${phone_b}"] = str(int(init_datas["${phone_a}"])+1)
            init_datas["${phone_c}"] = str(int(init_datas["${phone_a}"])+2)
            init_datas["${pre_phone_a}"] = str(int(init_datas["${phone_a}"])-3)

        return init_datas


    def modify_phone(self):
        res = int(self.sh_prepar_data.cell(row=2,column=2).value)
        self.sh_prepar_data.cell(row=2, column=2).value = str(res +3)


    def save_excelFile(self,):
        self.wb.save(self.excelpath)


    def get_caseDatas_all(self):
        all_case_datas = []
        for index in range(2,self.sh_case_data.max_row+1):
            case_data ={}

            case_data["case_id"] = self.sh_case_data.cell(row=index,column=1).value
            case_data["case_info"] = self.sh_case_data.cell(row=index,column=3).value
            case_data["api_name"] = self.sh_case_data.cell(row=index,column=4).value
            case_data["method"] = self.sh_case_data.cell(row=index, column=5).value
            case_data["url"] = self.sh_case_data.cell(row=index, column=6).value

            temp_case_data = self.sh_case_data.cell(row=index,column=7).value
            #获取初始值
            init_datas = self.get_init_datas()


            if temp_case_data is not None and len(init_datas)> 0:
                for key,value in init_datas.items():
                    #如果找到了则替换
                    if temp_case_data.find(key) != -1:
                        temp_case_data = temp_case_data.replace(str(key),str(value))

            case_data["request_data"] = temp_case_data
            case_data["expected_data"] = self.sh_case_data.cell(row=index, column=8).value
            case_data["compare_type"] = self.sh_case_data.cell(row=index, column=9).value
            # 判断本次用例是否需要对响应结果进行解析，并获取其中的值。
            if self.sh_case_data.cell(row=index, column=10).value is not None:
                case_data["related_exp"] = self.sh_case_data.cell(row=index, column=10).value

            if self.sh_case_data.cell(row=index, column=11).value is not None:
                case_data["related_exp1"] = self.sh_case_data.cell(row=index, column=11).value

            all_case_datas.append(case_data)
        return all_case_datas